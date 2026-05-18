"""Phase 14.11 - alarm rule bulk import parsing and validation.

Standalone module - no FastAPI/HTTP imports. The API layer wraps these
functions; smoke tests can call them directly.

Pipeline:
  1. parse_csv() / parse_xlsx() -> list[dict[str, Any]] of raw rows
  2. validate_rows(db, raw_rows) -> ImportResult with per-row outcomes
  3. (caller, if dry_run=false) commit_rows(db, validated) -> persists

Validation rules per row:
  - tag_name resolves to existing tag (case-sensitive name lookup)
  - rule_type exists in alarm_rule_types AND is_evaluable=true
    (frozen and any future taxonomy-only types are rejected with a
     clear message rather than a silent FK failure)
  - severity exists in alarm_severities
  - threshold is required, parses as float
  - deadband, on_delay_sec, off_delay_sec >= 0; default to 0 if blank
  - latched, enabled parse as bool; default false / true respectively
  - window_seconds REQUIRED for rate_of_change (1..86400); ignored
    for other rule types
  - (tag_id, rule_type) for {hi_hi, hi, lo, lo_lo} must not duplicate
    either an existing DB rule or another row in the same batch

The validator returns rich per-row results rather than throwing on
the first error. Dry-run UI shows the full picture before commit.
"""

from __future__ import annotations

import csv
import io
import logging
from dataclasses import dataclass, field
from typing import Any

from openpyxl import Workbook, load_workbook
from sqlalchemy import text
from sqlalchemy.orm import Session


log = logging.getLogger("alarms_import")


# Column names in the CSV/XLSX template. Order matters for the template
# output but not for parsing - parser is header-name-driven.
TEMPLATE_COLUMNS = [
    "tag_name", "rule_type", "severity", "threshold",
    "deadband", "on_delay_sec", "off_delay_sec",
    "latched", "window_seconds", "message_template", "enabled",
]

REQUIRED_COLUMNS = ["tag_name", "rule_type", "severity", "threshold"]

LEVEL_RULE_TYPES = ("hi_hi", "hi", "lo", "lo_lo")

# Two example rows shipped with the template so operators see the shape.
TEMPLATE_EXAMPLE_ROWS = [
    {
        "tag_name": "example_temp_inlet",
        "rule_type": "hi_hi",
        "severity": "critical",
        "threshold": "90.0",
        "deadband": "1.0",
        "on_delay_sec": "30",
        "off_delay_sec": "30",
        "latched": "true",
        "window_seconds": "",
        "message_template": "Inlet temperature critical: {value}",
        "enabled": "true",
    },
    {
        "tag_name": "example_flow_q1",
        "rule_type": "rate_of_change",
        "severity": "high",
        "threshold": "5.0",
        "deadband": "0.5",
        "on_delay_sec": "10",
        "off_delay_sec": "10",
        "latched": "false",
        "window_seconds": "60",
        "message_template": "Flow changing too fast",
        "enabled": "true",
    },
]


# ===========================================================================
# Result types
# ===========================================================================

@dataclass
class RowResult:
    row_number: int            # 1-based row in source file (header is row 1)
    tag_name: str
    rule_type: str
    severity: str
    threshold: float | None
    status: str                # 'ok' | 'error' | 'duplicate'
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    proposed: dict[str, Any] | None = None  # resolved insert payload

    def to_dict(self) -> dict[str, Any]:
        return {
            "row_number": self.row_number,
            "tag_name": self.tag_name,
            "rule_type": self.rule_type,
            "severity": self.severity,
            "threshold": self.threshold,
            "status": self.status,
            "errors": self.errors,
            "warnings": self.warnings,
            "proposed": self.proposed,
        }


@dataclass
class ImportSummary:
    total_rows: int
    ok_count: int
    error_count: int
    duplicate_count: int
    warning_count: int
    rows: list[RowResult]
    dry_run: bool
    committed: bool = False

    def to_dict(self) -> dict[str, Any]:
        return {
            "total_rows": self.total_rows,
            "ok_count": self.ok_count,
            "error_count": self.error_count,
            "duplicate_count": self.duplicate_count,
            "warning_count": self.warning_count,
            "rows": [r.to_dict() for r in self.rows],
            "dry_run": self.dry_run,
            "committed": self.committed,
        }


# ===========================================================================
# Parsing
# ===========================================================================

def parse_csv(content: bytes) -> list[dict[str, Any]]:
    """Parse CSV bytes into raw row dicts. Empty cells become empty
    strings; the validator normalises them to None later."""
    text_content = content.decode("utf-8-sig")  # tolerate BOM
    reader = csv.DictReader(io.StringIO(text_content))
    rows = []
    for raw in reader:
        # Strip whitespace from all cells
        cleaned = {(k or "").strip(): (v or "").strip() if v else ""
                   for k, v in raw.items()}
        rows.append(cleaned)
    return rows


def parse_xlsx(content: bytes) -> list[dict[str, Any]]:
    """Parse XLSX bytes via openpyxl. Reads the first sheet."""
    wb: Workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return []

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return []
    headers = [(h or "").strip() if isinstance(h, str) else str(h or "").strip()
               for h in header]

    rows = []
    for raw_row in rows_iter:
        # Skip fully-empty rows
        if all(c is None or (isinstance(c, str) and not c.strip())
               for c in raw_row):
            continue
        cells: dict[str, Any] = {}
        for hdr, val in zip(headers, raw_row):
            if not hdr:
                continue
            if isinstance(val, str):
                cells[hdr] = val.strip()
            elif val is None:
                cells[hdr] = ""
            else:
                cells[hdr] = val  # int, float, bool
        rows.append(cells)
    wb.close()
    return rows


# ===========================================================================
# Field coercion helpers
# ===========================================================================

def _as_bool(val: Any, default: bool) -> bool | None:
    """Returns True/False on recognized values, default on blank.
    Returns None if value is provided but unrecognized (caller flags error)."""
    if val is None or val == "":
        return default
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return bool(val)
    s = str(val).strip().lower()
    if s in ("true", "t", "yes", "y", "1"):
        return True
    if s in ("false", "f", "no", "n", "0"):
        return False
    return None


def _as_float(val: Any) -> float | None:
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).strip())
    except (ValueError, TypeError):
        return None


def _as_int(val: Any) -> int | None:
    if val is None or val == "":
        return None
    if isinstance(val, bool):
        # bool is an int subclass in Python; reject explicitly
        return None
    if isinstance(val, int):
        return val
    if isinstance(val, float):
        if val == int(val):
            return int(val)
        return None
    try:
        return int(str(val).strip())
    except (ValueError, TypeError):
        return None


# ===========================================================================
# Validation
# ===========================================================================

def validate_rows(db: Session, raw_rows: list[dict[str, Any]]) -> ImportSummary:
    """Validate every row against the live DB state. Returns a summary
    with per-row outcomes. Does not write anything to the DB."""
    if not raw_rows:
        return ImportSummary(
            total_rows=0, ok_count=0, error_count=0,
            duplicate_count=0, warning_count=0, rows=[],
            dry_run=True,
        )

    # ---- Pre-fetch reference data ----
    severity_codes = {
        r[0] for r in db.execute(
            text("SELECT code FROM alarm_severities")
        ).fetchall()
    }
    rule_type_rows = db.execute(text("""
        SELECT code, is_evaluable FROM alarm_rule_types
    """)).fetchall()
    rule_type_evaluable: dict[str, bool] = {r[0]: r[1] for r in rule_type_rows}

    # Tag name -> id, fetched in one query.
    tag_names = {(r.get("tag_name") or "").strip() for r in raw_rows
                 if r.get("tag_name")}
    tag_rows = db.execute(
        text("SELECT id, name FROM tags WHERE name = ANY(:names)"),
        {"names": list(tag_names)},
    ).fetchall()
    tag_id_by_name: dict[str, int] = {r[1]: r[0] for r in tag_rows}

    # Existing level rules in DB (for duplicate detection on level types).
    if tag_id_by_name:
        existing_rows = db.execute(text("""
            SELECT tag_id, rule_type FROM alarm_rules
            WHERE rule_type = ANY(:types) AND tag_id = ANY(:ids)
        """), {
            "types": list(LEVEL_RULE_TYPES),
            "ids": list(tag_id_by_name.values()),
        }).fetchall()
        existing_level_rules = {(r[0], r[1]) for r in existing_rows}
    else:
        existing_level_rules = set()

    # ---- Per-row validation ----
    results: list[RowResult] = []
    batch_level_keys: set[tuple[int, str]] = set()  # (tag_id, rule_type)

    for idx, raw in enumerate(raw_rows, start=2):  # row 1 is header
        result = _validate_one_row(
            raw, idx,
            severity_codes=severity_codes,
            rule_type_evaluable=rule_type_evaluable,
            tag_id_by_name=tag_id_by_name,
            existing_level_rules=existing_level_rules,
            batch_level_keys=batch_level_keys,
        )
        # Add to batch tracking if this row is OK and it's a level type
        if (result.status == "ok" and result.proposed
                and result.proposed["rule_type"] in LEVEL_RULE_TYPES):
            key = (result.proposed["tag_id"], result.proposed["rule_type"])
            batch_level_keys.add(key)
        results.append(result)

    ok = sum(1 for r in results if r.status == "ok")
    err = sum(1 for r in results if r.status == "error")
    dup = sum(1 for r in results if r.status == "duplicate")
    warn = sum(1 for r in results if r.warnings)

    return ImportSummary(
        total_rows=len(results),
        ok_count=ok,
        error_count=err,
        duplicate_count=dup,
        warning_count=warn,
        rows=results,
        dry_run=True,
    )


def _validate_one_row(
    raw: dict[str, Any],
    row_number: int,
    *,
    severity_codes: set[str],
    rule_type_evaluable: dict[str, bool],
    tag_id_by_name: dict[str, int],
    existing_level_rules: set[tuple[int, str]],
    batch_level_keys: set[tuple[int, str]],
) -> RowResult:
    errors: list[str] = []
    warnings: list[str] = []

    tag_name = str(raw.get("tag_name", "") or "").strip()
    rule_type = str(raw.get("rule_type", "") or "").strip()
    severity = str(raw.get("severity", "") or "").strip()
    threshold_raw = raw.get("threshold", "")

    threshold = _as_float(threshold_raw)

    # ---- Required-presence checks ----
    if not tag_name:
        errors.append("tag_name is required")
    if not rule_type:
        errors.append("rule_type is required")
    if not severity:
        errors.append("severity is required")
    if threshold_raw == "" or threshold_raw is None:
        errors.append("threshold is required")
    elif threshold is None:
        errors.append(f"threshold must be a number, got {threshold_raw!r}")

    # ---- Reference-data checks ----
    tag_id: int | None = None
    if tag_name:
        tag_id = tag_id_by_name.get(tag_name)
        if tag_id is None:
            errors.append(f"tag '{tag_name}' does not exist")

    if rule_type:
        if rule_type not in rule_type_evaluable:
            errors.append(
                f"rule_type '{rule_type}' is not a recognized type "
                f"(known: {sorted(rule_type_evaluable.keys())})"
            )
        elif not rule_type_evaluable[rule_type]:
            errors.append(
                f"rule_type '{rule_type}' is taxonomy-only "
                f"(is_evaluable=false) and cannot be assigned to a rule"
            )

    if severity and severity not in severity_codes:
        errors.append(
            f"severity '{severity}' is not recognized "
            f"(known: {sorted(severity_codes)})"
        )

    # ---- Numeric / boolean coercion ----
    deadband = _as_float(raw.get("deadband", "")) if raw.get("deadband", "") != "" else 0.0
    if deadband is None:
        errors.append(f"deadband must be a number, got {raw.get('deadband')!r}")
    elif deadband < 0:
        errors.append(f"deadband must be >= 0, got {deadband}")

    on_delay = _as_int(raw.get("on_delay_sec", "")) if raw.get("on_delay_sec", "") != "" else 0
    if on_delay is None:
        errors.append(f"on_delay_sec must be an integer, got {raw.get('on_delay_sec')!r}")
    elif on_delay < 0:
        errors.append(f"on_delay_sec must be >= 0, got {on_delay}")

    off_delay = _as_int(raw.get("off_delay_sec", "")) if raw.get("off_delay_sec", "") != "" else 0
    if off_delay is None:
        errors.append(f"off_delay_sec must be an integer, got {raw.get('off_delay_sec')!r}")
    elif off_delay < 0:
        errors.append(f"off_delay_sec must be >= 0, got {off_delay}")

    latched = _as_bool(raw.get("latched", ""), default=False)
    if latched is None:
        errors.append(f"latched must be true/false, got {raw.get('latched')!r}")

    enabled = _as_bool(raw.get("enabled", ""), default=True)
    if enabled is None:
        errors.append(f"enabled must be true/false, got {raw.get('enabled')!r}")

    # ---- window_seconds: required for rate_of_change and frozen ----
    # Both rule types use a window to compute their metric. The
    # worker tolerates missing window for deviation (falls back to
    # DEFAULT_WINDOW_SECONDS=60) but for rate_of_change and frozen
    # the operator should make the window explicit. Phase 14.10.
    window_seconds: int | None = None
    window_raw = raw.get("window_seconds", "")
    WINDOWED_REQUIRED = ("rate_of_change", "frozen")
    if rule_type in WINDOWED_REQUIRED:
        if window_raw == "" or window_raw is None:
            errors.append(
                f"window_seconds is required for {rule_type} rules"
            )
        else:
            window_seconds = _as_int(window_raw)
            if window_seconds is None:
                errors.append(
                    f"window_seconds must be an integer, got {window_raw!r}"
                )
            elif window_seconds < 1 or window_seconds > 86400:
                errors.append(
                    f"window_seconds must be in [1, 86400], got {window_seconds}"
                )
    elif window_raw != "" and window_raw is not None:
        # Other rule types: window_seconds ignored. Warn rather than
        # error so accidental column carry-over doesn't block the
        # import. (deviation also uses the column but tolerates absence.)
        warnings.append(
            f"window_seconds is only used for rate_of_change / frozen / "
            f"deviation rules; value {window_raw!r} ignored for "
            f"rule_type '{rule_type}'"
        )

    message_template = str(raw.get("message_template", "") or "").strip()
    if len(message_template) > 500:
        errors.append(
            f"message_template too long ({len(message_template)} chars; max 500)"
        )

    # ---- Duplicate check for level types ----
    status = "ok"
    if (rule_type in LEVEL_RULE_TYPES and tag_id is not None and not errors):
        key = (tag_id, rule_type)
        if key in existing_level_rules:
            status = "duplicate"
            errors.append(
                f"a {rule_type} rule already exists in the database for tag '{tag_name}'"
            )
        elif key in batch_level_keys:
            status = "duplicate"
            errors.append(
                f"this batch already contains another {rule_type} rule for tag '{tag_name}'"
            )

    if errors and status == "ok":
        status = "error"

    proposed: dict[str, Any] | None = None
    if status == "ok":
        proposed = {
            "tag_id": tag_id,
            "tag_name": tag_name,
            "rule_type": rule_type,
            "severity": severity,
            "threshold": threshold,
            "deadband": deadband,
            "on_delay_sec": on_delay,
            "off_delay_sec": off_delay,
            "latched": latched,
            "enabled": enabled,
            "window_seconds": window_seconds,
            "message_template": message_template or None,
        }

    return RowResult(
        row_number=row_number,
        tag_name=tag_name,
        rule_type=rule_type,
        severity=severity,
        threshold=threshold,
        status=status,
        errors=errors,
        warnings=warnings,
        proposed=proposed,
    )


# ===========================================================================
# Commit
# ===========================================================================

def commit_rows(db: Session, summary: ImportSummary) -> None:
    """Insert all OK rows in a single transaction. Caller is
    responsible for db.commit() / db.rollback().

    Raises if any insert fails. The caller's transaction context
    handles rollback so the whole batch is atomic.
    """
    inserts = [r.proposed for r in summary.rows if r.status == "ok" and r.proposed]
    if not inserts:
        return

    for p in inserts:
        db.execute(text("""
            INSERT INTO alarm_rules (
                tag_id, rule_type, severity, threshold,
                deadband, on_delay_sec, off_delay_sec,
                latched, enabled, window_seconds, message_template
            ) VALUES (
                :tag_id, :rule_type, :severity, :threshold,
                :deadband, :on_delay_sec, :off_delay_sec,
                :latched, :enabled, :window_seconds, :message_template
            )
        """), p)


# ===========================================================================
# Template generation
# ===========================================================================

def render_template_csv() -> bytes:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=TEMPLATE_COLUMNS,
                            lineterminator="\n")
    writer.writeheader()
    for row in TEMPLATE_EXAMPLE_ROWS:
        writer.writerow(row)
    return buf.getvalue().encode("utf-8")


def render_template_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "alarm_rules"
    ws.append(TEMPLATE_COLUMNS)
    for row in TEMPLATE_EXAMPLE_ROWS:
        ws.append([row.get(c, "") for c in TEMPLATE_COLUMNS])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
